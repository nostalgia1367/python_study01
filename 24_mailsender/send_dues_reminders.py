import openpyxl
import smtplib
from email.message import EmailMessage
import sys


def check_unpaid_memebers():
    # 엑셀 파일을 연다.
    wb = openpyxl.load_workbook('members.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')

    unpaid_members = {}
    # 회원 회비납부 여부 확인
    for r in range(2, sheet.max_row + 1):
        payment = sheet.cell(row=r, column=sheet.max_column).value
        if payment != 'paid':
            name = sheet.cell(row=r, column=1).value
            email = sheet.cell(row=r, column=2).value
            unpaid_members[name] = email

    return unpaid_members


def main():
    # 회비 미납회원 명단 확보
    unpaid_members = check_unpaid_memebers()

    # Log in to email account.
    smtp = smtplib.SMTP('smtp.gmail.com', 587)
    smtp.ehlo()
    smtp.starttls()
    smtp.login('my_email_address@gmail.com', 'password')

    # 미납회원에게 리마인드 메일을 보낸다.
    for name, email in unpaid_members.items():
        msg = EmailMessage()
        msg['Subject'] = '회비 미납 안내'
        msg['From'] = ''
        msg['To'] = email
        msg.set_content('''
        안녕하세요 {}님.

        이번달 회비가 미납되어 알려드립니다.
        빠른 시일내에 납부 부탁드립니다.
        '''.format(name))

        print('Sending email to {}...'.format(email))
        sendmail_status = smtp.send_message(msg)

        if sendmail_status != {}:
            print('{} 에게 메일 발송이 실패했습니다.: {}'.format(email, sendmail_status))

    smtp.quit()


if __name__=='__main__':
    main()
