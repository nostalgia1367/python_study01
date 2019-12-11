#! python3
# updateProduce.py - Corrects costs in produce sales spreadsheet.

"""
엑셀 지원 파이썬 모듈
    xlwt
    OpenPyXL    http://openpyxl.readthedocs.org/
    XlsxWriter
    PyExcelerate

openpyxl 모듈
    - 로컬에 엑셀 프로그램이 설치되지 않더라도 엑셀 엑성과 읽기 가능
    - 대용량 지원, 이미지 지원
    - pip install openpyxl


CSV : Comma Separated Value
콤마로 구분된 데이터를 라인단ㄷ위로 읽어 리스트나 딕셔너리 같은 자료구조
파일의 운영체 줄바꿈 형태
리눅스 : \n
윈도우 : \r\n


엑셀과 CSV 차이
값의 데이터타입이 없다(모든 값이 문자열)
각셀 폰트나 컬러 없음
이미지나 차트 포함 불가
셀의 높이나 너비 없다
워크시트가 없다(하나의 워크시트만 존재)

"""

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

# The produce types and their updated prices
PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}

# Loop through the rows and update the prices.
for rowNum in range(2, sheet.get_highest_row()): # skip the first row
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('updatedProduceSales.xlsx')
